# -*- coding: utf-8 -*-
"""
数据对齐：把原始数据文件解析并按 (node, date, hour) 合并为 master 长表。

输入（项目根目录，相对本文件 ../）：
  价格数据/*.xlsx        —— 只用非 -c 文件；Date 向下填充；B 列无表头，DA/RTPD/DARTPD Return
  load_CA_ISO_TAC_2DA.csv    —— 官方日前负荷预测（未来已知特征）
  load_CA_ISO_TAC_ACTUAL.csv —— 实际负荷（仅历史）
  zone_weather_hourly.csv    —— 逐小时天气（zone: NP15/SP15/ZP26）
  节点位置.xlsx              —— node -> zone 映射

输出：
  code/data/master.csv       —— 长表: node, date, hour, da_price, rtpd_price, spread,
                                zone, load_2da, load_actual, t2m, ssrd, wind100
"""
import os
import glob
import openpyxl
import pandas as pd
import numpy as np

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "code", "data")
PRICE_DIR = os.path.join(ROOT, "价格数据")
OUT = os.path.join(DATA, "master.csv")

HOUR_COLS = ["H%d" % h for h in range(1, 25)]


def read_price_node(path, node):
    """读取单个节点价格 xlsx，返回 [(node, date, hour, market, value)]。"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Sheet1"]
    out = []
    cur_date = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        # row[0]=Date, row[1]=market(无表头), row[2]=Hours, row[3]=Avg,
        # row[4]=Total, row[5..28]=H1..H24
        if row[0] is not None:
            cur_date = row[0]
        market = row[1]
        if cur_date is None or market is None:
            continue
        for h in range(1, 25):
            val = row[4 + h]  # row[5]=H1 -> row[28]=H24
            if val is None:
                continue
            out.append((node, pd.Timestamp(cur_date).date(), h, market, float(val)))
    return out


def read_all_prices(nodes):
    """读取所有节点价格，返回 DataFrame: node,date,hour,da_price,rtpd_price,spread"""
    records = []
    for node in nodes:
        for path in glob.glob(os.path.join(PRICE_DIR, "*.xlsx")):
            base = os.path.basename(path)
            if base.endswith("-c.xlsx"):
                continue  # 组合价格文件，业务明确不用
            if base.startswith(node + ".") or base.startswith(node + "-"):
                # 匹配 "SNLNDRO_1_N001.xlsx"
                if base.replace(".xlsx", "") == node:
                    records.extend(read_price_node(path, node))
                    break
        else:
            raise FileNotFoundError("未找到节点 %s 的价格文件" % node)

    df = pd.DataFrame(records, columns=["node", "date", "hour", "market", "value"])
    wide = df.pivot_table(index=["node", "date", "hour"], columns="market", values="value").reset_index()
    wide.columns.name = None
    # 各 market 可能缺失 -> 用列存在性兜底
    da = wide["DA"] if "DA" in wide.columns else np.nan
    rt = wide["RTPD"] if "RTPD" in wide.columns else np.nan
    wide["da_price"] = da
    wide["rtpd_price"] = rt
    wide["spread"] = wide["da_price"] - wide["rtpd_price"]
    return wide[["node", "date", "hour", "da_price", "rtpd_price", "spread"]]


def read_load_2da():
    """日前负荷预测：Date 格式 YYYY/M/D（不补零）。"""
    path = os.path.join(ROOT, "load_CA_ISO_TAC_2DA.csv")
    df = pd.read_csv(path)
    df["date"] = pd.to_datetime(df["Date"].astype(str), format="mixed").dt.date
    df = df.melt(id_vars="date", value_vars=HOUR_COLS, var_name="hour_col", value_name="load_2da")
    df["hour"] = df["hour_col"].str.extract(r"(\d+)").astype(int)
    return df[["date", "hour", "load_2da"]]


def read_load_actual():
    """实际负荷：Date 格式 YYYY-MM-DD。"""
    path = os.path.join(ROOT, "load_CA_ISO_TAC_ACTUAL.csv")
    df = pd.read_csv(path)
    df["date"] = pd.to_datetime(df["Date"]).dt.date
    df = df.melt(id_vars="date", value_vars=HOUR_COLS, var_name="hour_col", value_name="load_actual")
    df["hour"] = df["hour_col"].str.extract(r"(\d+)").astype(int)
    return df[["date", "hour", "load_actual"]]


def read_weather():
    """天气：zone, valid_pt(YYYY-MM-DD HH:MM:SS), t2m_c, ssrd_wm2, wind100, Date。
    假设 valid_pt 为该小时起始：00:00 -> H1 ... 23:00 -> H24。
    """
    path = os.path.join(ROOT, "zone_weather_hourly.csv")
    df = pd.read_csv(path)
    vt = pd.to_datetime(df["valid_pt"])
    df["date"] = vt.dt.date
    df["hour"] = vt.dt.hour + 1  # 0:00 -> H1
    return df[["zone", "date", "hour", "t2m_c", "ssrd_wm2", "wind100"]].rename(
        columns={"t2m_c": "t2m", "ssrd_wm2": "ssrd", "wind100": "wind100"}
    )


def read_node_zone():
    """节点位置：node -> zone 映射。"""
    path = os.path.join(ROOT, "节点位置.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Sheet1"]
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        mapping[str(row[0]).strip()] = str(row[3]).strip()  # 区域(Zone)
    return mapping


def build_master(nodes):
    prices = read_all_prices(nodes)
    load2 = read_load_2da()
    loada = read_load_actual()
    weather = read_weather()
    node_zone = read_node_zone()

    prices["zone"] = prices["node"].map(node_zone)
    # 节点关联 zone 天气 + 系统负荷
    master = prices.merge(load2, on=["date", "hour"], how="left")
    master = master.merge(loada, on=["date", "hour"], how="left")
    master = master.merge(weather, on=["zone", "date", "hour"], how="left")

    cols = ["node", "date", "hour", "da_price", "rtpd_price", "spread", "zone",
            "load_2da", "load_actual", "t2m", "ssrd", "wind100"]
    master = master[cols].sort_values(["node", "date", "hour"]).reset_index(drop=True)
    return master


if __name__ == "__main__":
    os.makedirs(DATA, exist_ok=True)
    nodes = ["SNLNDRO_1_N001", "CONTROLX_1_N001", "ELCAJNGT_7_N001"]
    m = build_master(nodes)
    m.to_csv(OUT, index=False)
    print("rows:", len(m))
    print("date range:", m["date"].min(), "->", m["date"].max())
    for n in nodes:
        sub = m[m["node"] == n]
        print(n, "rows:", len(sub), "dates:", sub["date"].min(), "->", sub["date"].max())
    print("columns:", list(m.columns))
    print("NaN counts:\n", m.isna().sum()[m.isna().sum() > 0])
    print("sample:\n", m[m["date"] == m["date"].min()].head(6).to_string())
